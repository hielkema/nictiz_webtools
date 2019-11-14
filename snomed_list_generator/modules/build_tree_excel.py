from __future__ import division
import json, requests, re, time, sys, pandas, pprint, pickle, datetime
from array import array
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from .snowstorm_client import Snowstorm

# start loopteller
loopnummer = 0

# timestamp voor bestand
ts = time.time()
st = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y %H:%M:%S')

# Create snowstorm client object
snowstorm = Snowstorm(baseUrl="https://snowstorm.test-nictiz.nl", defaultBranchPath="MAIN/SNOMEDCT-NL", preferredLanguage="nl")

# Map voor schrijven van output
writeFolder = "/webserver/static_files/tree/"
if __name__ == "__main__":
    writeFolder = ""
    
    
def children_array(concept, getParents=False):
    # Create lists for children and parents SCTIDs
    _children = []
    _parents = []
    
    # Get all children of the current concept
    _children_local = snowstorm.getChildren(id=concept)
    # print(_children_local)
    
    # Get parents, of if empty create empty list
    if getParents:
        try:
            _parents_local = snowstorm.getParents(id=concept)
        except:
            _parents_local = []
        
    # Put only the children SCTIDs in temporary list
    try:
        for value in _children_local:
            try:
                _children.append(value)
            except:
                error=True
    except:
        error=True
    # Put only the parent SCTIDs in temporary list
    if getParents:
        for value in _parents_local:
            try:
                _parents.append(value)
            except:
                error=True
    
    # Add all parent concepts to the concept dict, as these will not be retrieved in the recursive loop.
    for value in _parents:
        _current = snowstorm.getDescriptions(id=value)
        nl_fsn = _current['categorized'].get('31000146106',{}).get('fsn',{}).get('term',"Niet vertaald")
        en_fsn = _current['categorized'].get('900000000000509007').get('fsn',{}).get('term',"Niet vertaald")
        concepts.update({
            value : {
                'conceptId'     :   value,
                'FSE'           :   en_fsn
            } 
        })
    
    # Add SNOMED data of current concept to the concept dict
    _current = snowstorm.getDescriptions(id=concept)
    nl_fsn = _current['categorized'].get('31000146106',{}).get('fsn',{}).get('term',"Niet vertaald")
    en_fsn = _current['categorized'].get('900000000000509007').get('fsn',{}).get('term',"Niet vertaald")
    concepts.update({
            concept : {
                'conceptId'     :   concept,
                'FSE'           :   en_fsn,
                'FSN'           :   nl_fsn
            } 
        })
    # Add relations of current concept to the relations dict from temporary lists created earlier
    relations.update({
            concept : { 
                'children'  :   _children,
                'parents'   :   _parents
                }
        })
    # Recursive loop; repeat for all children of the current concept
    # Not for all parents; would create infinite loop
    for value in _children:
        children_array(value)
    
    return [concepts, relations]

def excel_tree(concept):
    global concepts, relations, row, column, level

    snowstorm.loadCache()


    # Extract data from snowstorm
    # Optional: Retrieve parents in addition to children - approx. 15% performance hit.
    concepts = {}
    relations = {}
    tree = children_array(concept, getParents=False)
    
    concepts = tree[0]
    relations = tree[1]
    
    time_extraction = time.time()
    runtime_extraction = round(time_extraction - ts, 0)
    print("Execution time data extraction: {}seconds.".format(runtime_extraction))


    wb = Workbook()
    timestamp = str(round(ts))
    dest_filename = '{} - {}.xlsx'.format(str(concept), timestamp)
    ws1 = wb.active
    ws1.title = "Flat Snomed overview"

    row = 3
    column = 1
    def writeChildren(concept, level):
        global row
        level += 1
        for value in relations[str(concept)]['children']:
            ws1.cell(row=row, column=column).value = value
            ws1.cell(row=row, column=column+1).value = concepts[str(value)]['FSE']
            try:
                ws1.cell(row=row, column=column+2).value = concepts[str(value)]['FSN']
            except:
                ws1.cell(row=row, column=column+2).value = "Niet vertaald"
            ws1.cell(row=row, column=column+3).value = level
            row+=1
            writeChildren(value, level)

    # Write titles
    ws1.cell(row=1, column=1).value = "Snomed ID"
    ws1.cell(row=1, column=2).value = "FSN Engels"
    ws1.cell(row=1, column=3).value = "FSN Nederlands"
    ws1.cell(row=1, column=4).value = "Relatie niveau"
    # Write focus concept    
    ws1.cell(row=2, column=1).value = concept
    ws1.cell(row=2, column=2).value = concepts[str(concept)]['FSE']
    try:
        ws1.cell(row=2, column=3).value = concepts[str(concept)]['FSN']
    except:
        ws1.cell(row=2, column=3).value = ""
    ws1.cell(row=2, column=4).value = 0
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 80
    ws1.column_dimensions['C'].width = 80
    ws1.column_dimensions['D'].width = 10

    # Children of focus concept, start at level 0
    writeChildren(concept, 0)    

    wb.save(filename = writeFolder+dest_filename)

    time_complete = time.time()
    runtime_export = round(time_complete - time_extraction, 0)
    runtime_total = round(time_complete - ts, 0)
    print("Execution time data export: {}seconds.\nTotal execution time: {}seconds.".format(runtime_export, runtime_total))
    
    # Write snowstorm cache
    snowstorm.writeCache()
    
    return [dest_filename, concepts[str(concept)]['FSE']]
    
if __name__ == "__main__":
    excel_tree("74400008") #74400008 / 87290003
